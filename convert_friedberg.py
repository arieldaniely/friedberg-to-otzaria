from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

import openpyxl


WILNA_NAME = "וילנא"
SPECIAL_FILE_MARKER = "קובץ מיוחד של סינופסיס"
INTRO_DIR_NAME = "00 - מבואות פרידברג"
WHITESPACE_RE = re.compile(r"\s+")
TOKEN_RE = re.compile(r"\S+")
TRACTATE_DIR_RE = re.compile(r"^(?P<order>\d+)\s*-\s*(?P<name>.+)$")
REF_PART_RE = re.compile(r'^(?P<amud>ע["״][אב])\s+(?P<lines>\d+(?:\s*-\s*\d+)?)$')

HEBREW_LETTER_VALUES = {
    "א": 1,
    "ב": 2,
    "ג": 3,
    "ד": 4,
    "ה": 5,
    "ו": 6,
    "ז": 7,
    "ח": 8,
    "ט": 9,
    "י": 10,
    "כ": 20,
    "ך": 20,
    "ל": 30,
    "מ": 40,
    "ם": 40,
    "נ": 50,
    "ן": 50,
    "ס": 60,
    "ע": 70,
    "פ": 80,
    "ף": 80,
    "צ": 90,
    "ץ": 90,
    "ק": 100,
    "ר": 200,
    "ש": 300,
    "ת": 400,
}


@dataclass(frozen=True)
class RefPart:
    amud: str
    line_start: int
    line_end: int
    label: str

    @property
    def line_count(self) -> int:
        return self.line_end - self.line_start + 1


@dataclass
class Segment:
    daf: int
    amud: str
    ref_label: str
    start_line: int
    order: int
    witness_names: list[str]
    texts: list[str]
    word_rows: list[list[str]]
    vilna_index: int | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert Friedberg Excel synopses into tractate text files."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("פרידברג"),
        help="Source directory that contains the tractate folders.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output"),
        help="Output directory for the generated text files.",
    )
    parser.add_argument(
        "--tractate",
        type=str,
        default=None,
        help="Process only tractates whose normalized name contains this value.",
    )
    return parser.parse_args()


def warn(message: str) -> None:
    print(f"[warn] {message}", file=sys.stderr)


def normalize_spaces(text: str) -> str:
    return WHITESPACE_RE.sub(" ", text).strip()


def strip_control_characters(text: str) -> str:
    kept: list[str] = []
    for char in text:
        if char in "\t\n\r":
            kept.append(" ")
            continue
        category = unicodedata.category(char)
        if category == "Cf":
            continue
        kept.append(char)
    return "".join(kept)


def clean_cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = strip_control_characters(text)
    return normalize_spaces(text)


def clean_ref_text(value: object) -> str:
    text = clean_cell_text(value)
    if not text:
        return ""
    return text.replace(" ;", ";").replace("; ", "; ").strip()


def normalize_compare_token(token: str) -> str:
    token = clean_cell_text(token)
    token = token.replace("״", '"').replace("׳", "'")
    return token


def tokenize_with_spans(text: str) -> tuple[list[str], list[tuple[int, int]]]:
    tokens: list[str] = []
    spans: list[tuple[int, int]] = []
    for match in TOKEN_RE.finditer(text):
        tokens.append(match.group(0))
        spans.append(match.span())
    return tokens, spans


def wrap_token_indices(text: str, token_indices: set[int]) -> str:
    if not token_indices:
        return text

    matches = list(TOKEN_RE.finditer(text))
    if not matches:
        return text

    ordered = sorted(index for index in token_indices if 0 <= index < len(matches))
    if not ordered:
        return text

    ranges: list[tuple[int, int]] = []
    range_start = ordered[0]
    range_end = ordered[0]
    for index in ordered[1:]:
        if index == range_end + 1:
            range_end = index
            continue
        ranges.append((range_start, range_end))
        range_start = range_end = index
    ranges.append((range_start, range_end))

    pieces: list[str] = []
    last_end = 0
    for start_index, end_index in ranges:
        start_char = matches[start_index].start()
        end_char = matches[end_index].end()
        pieces.append(text[last_end:start_char])
        pieces.append(f"<b><i>{text[start_char:end_char]}</i></b>")
        last_end = end_char
    pieces.append(text[last_end:])
    return "".join(pieces)


def apply_diff_highlights(text: str, vilna_text: str) -> str:
    text_tokens, _ = tokenize_with_spans(text)
    vilna_tokens, _ = tokenize_with_spans(vilna_text)
    if not text_tokens:
        return text

    matcher = SequenceMatcher(
        a=[normalize_compare_token(token) for token in vilna_tokens],
        b=[normalize_compare_token(token) for token in text_tokens],
        autojunk=False,
    )
    highlighted: set[int] = set()
    for opcode, _a1, _a2, b1, b2 in matcher.get_opcodes():
        if opcode in {"replace", "insert"}:
            highlighted.update(range(b1, b2))
    return wrap_token_indices(text, highlighted)


def map_word_tokens_to_text_tokens(
    word_tokens: list[str], text_tokens: list[str]
) -> dict[int, int] | None:
    if not word_tokens or not text_tokens:
        return {} if not word_tokens else None

    if len(word_tokens) == len(text_tokens):
        return {index: index for index in range(len(word_tokens))}

    matcher = SequenceMatcher(
        a=[normalize_compare_token(token) for token in word_tokens],
        b=[normalize_compare_token(token) for token in text_tokens],
        autojunk=False,
    )
    mapping: dict[int, int] = {}
    for opcode, a1, a2, b1, b2 in matcher.get_opcodes():
        if opcode == "equal":
            for offset in range(a2 - a1):
                mapping[a1 + offset] = b1 + offset
        elif opcode == "replace":
            shared = min(a2 - a1, b2 - b1)
            for offset in range(shared):
                mapping[a1 + offset] = b1 + offset

    coverage = len(mapping) / max(len(word_tokens), 1)
    if coverage < 0.6:
        return None
    return mapping


def apply_word_sheet_highlights(
    text: str,
    word_tokens: list[str],
    changed_positions: list[int],
) -> str | None:
    text_tokens, _ = tokenize_with_spans(text)
    if not text_tokens:
        return text

    mapping = map_word_tokens_to_text_tokens(word_tokens, text_tokens)
    if mapping is None:
        return None

    highlighted: set[int] = set()
    for word_position in changed_positions:
        text_position = mapping.get(word_position)
        if text_position is not None:
            highlighted.add(text_position)
    return wrap_token_indices(text, highlighted)


def parse_ref_part(ref_text: str) -> RefPart | None:
    ref_text = normalize_spaces(ref_text)
    match = REF_PART_RE.match(ref_text)
    if not match:
        return None

    line_text = match.group("lines")
    if "-" in line_text:
        left, right = [part.strip() for part in line_text.split("-", 1)]
        line_start = int(left)
        line_end = int(right)
    else:
        line_start = line_end = int(line_text)

    return RefPart(
        amud=match.group("amud"),
        line_start=line_start,
        line_end=line_end,
        label=f"{match.group('amud')} {line_text}",
    )


def hebrew_to_int(text: str) -> int:
    total = 0
    for char in clean_cell_text(text):
        if char in {'"', "״", "'", "׳"}:
            continue
        try:
            total += HEBREW_LETTER_VALUES[char]
        except KeyError as exc:
            raise ValueError(f"Unsupported Hebrew numeral: {text!r}") from exc
    if total <= 0:
        raise ValueError(f"Could not parse Hebrew numeral: {text!r}")
    return total


def int_to_hebrew(number: int) -> str:
    if number <= 0:
        raise ValueError(f"Unsupported daf number: {number}")

    parts: list[str] = []
    hundreds = [
        (400, "ת"),
        (300, "ש"),
        (200, "ר"),
        (100, "ק"),
    ]
    tens = [
        (90, "צ"),
        (80, "פ"),
        (70, "ע"),
        (60, "ס"),
        (50, "נ"),
        (40, "מ"),
        (30, "ל"),
        (20, "כ"),
        (10, "י"),
    ]
    ones = [
        (9, "ט"),
        (8, "ח"),
        (7, "ז"),
        (6, "ו"),
        (5, "ה"),
        (4, "ד"),
        (3, "ג"),
        (2, "ב"),
        (1, "א"),
    ]

    remainder = number
    while remainder >= 400:
        parts.append("ת")
        remainder -= 400
    for value, letter in hundreds[1:]:
        while remainder >= value:
            parts.append(letter)
            remainder -= value

    if remainder == 15:
        parts.append("טו")
        remainder = 0
    elif remainder == 16:
        parts.append("טז")
        remainder = 0

    for value, letter in tens:
        while remainder >= value:
            parts.append(letter)
            remainder -= value
    for value, letter in ones:
        while remainder >= value:
            parts.append(letter)
            remainder -= value

    return "".join(parts)


def parse_tractate_dir(path: Path) -> tuple[int, str] | None:
    match = TRACTATE_DIR_RE.match(path.name)
    if not match:
        return None
    return int(match.group("order")), match.group("name")


def normalize_tractate_name(dir_name: str) -> str:
    name = dir_name.strip()
    if name.startswith("מסכת "):
        name = name[5:].strip()
    return name


def extract_daf_number_from_filename(path: Path) -> int:
    stem = path.stem
    if "," not in stem:
        raise ValueError(f"Workbook file name does not contain daf marker: {path.name}")
    daf_text = stem.rsplit(",", 1)[1].strip()
    return hebrew_to_int(daf_text)


def choose_sheet_name(sheet_names: Iterable[str], keyword: str, fallback_index: int) -> str:
    sheet_names = list(sheet_names)
    for sheet_name in sheet_names:
        if keyword in sheet_name:
            return sheet_name
    return sheet_names[fallback_index]


def get_witness_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[tuple[int, str]]:
    header_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    witness_columns: list[tuple[int, str]] = []
    for column_index, value in enumerate(header_row[1:], start=2):
        witness_name = clean_cell_text(value)
        if witness_name:
            witness_columns.append((column_index, witness_name))
    return witness_columns


def read_word_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    witness_columns: list[tuple[int, str]],
) -> dict[str, list[list[str]]]:
    grouped_rows: dict[str, list[list[str]]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        ref = clean_ref_text(row[0] if row else None)
        if not ref:
            continue

        token_row: list[str] = []
        for column_index, _name in witness_columns:
            raw_value = row[column_index - 1] if column_index - 1 < len(row) else None
            token_row.append(clean_cell_text(raw_value))
        grouped_rows.setdefault(ref, []).append(token_row)
    return grouped_rows


def count_nonempty_tokens(word_rows: list[list[str]], witness_index: int) -> int:
    return sum(1 for row in word_rows if witness_index < len(row) and row[witness_index])


def allocate_units(total_units: int, weights: list[int]) -> list[int]:
    if not weights:
        return []
    if total_units <= 0:
        return [0] * len(weights)

    weight_sum = sum(weights)
    if weight_sum <= 0:
        weight_sum = len(weights)
        weights = [1] * len(weights)

    raw_allocations = [total_units * weight / weight_sum for weight in weights]
    unit_counts = [int(value) for value in raw_allocations]
    remainder = total_units - sum(unit_counts)

    ranked = sorted(
        range(len(weights)),
        key=lambda index: (raw_allocations[index] - unit_counts[index], -index),
        reverse=True,
    )
    for index in ranked[:remainder]:
        unit_counts[index] += 1
    return unit_counts


def split_rows_by_boundaries(word_rows: list[list[str]], boundaries: list[int]) -> list[list[list[str]]]:
    slices: list[list[list[str]]] = []
    start = 0
    for end in boundaries:
        slices.append(word_rows[start:end])
        start = end
    slices.append(word_rows[start:])
    return slices


def choose_split_boundaries(
    word_rows: list[list[str]],
    vilna_index: int | None,
    line_counts: list[int],
) -> list[int]:
    if len(line_counts) <= 1:
        return []

    if vilna_index is not None:
        vilna_positions = [
            index for index, row in enumerate(word_rows) if vilna_index < len(row) and row[vilna_index]
        ]
    else:
        vilna_positions = []

    if vilna_positions:
        token_counts = allocate_units(len(vilna_positions), line_counts)
        boundaries: list[int] = []
        cumulative_tokens = 0
        for count in token_counts[:-1]:
            cumulative_tokens += count
            if cumulative_tokens <= 0:
                boundaries.append(0)
            elif cumulative_tokens >= len(vilna_positions):
                boundaries.append(len(word_rows))
            else:
                boundaries.append(vilna_positions[cumulative_tokens - 1] + 1)
        return boundaries

    row_counts = allocate_units(len(word_rows), line_counts)
    boundaries = []
    cumulative_rows = 0
    for count in row_counts[:-1]:
        cumulative_rows += count
        boundaries.append(cumulative_rows)
    return boundaries


def split_text_by_token_counts(
    text: str,
    part_word_counts: list[int],
    line_counts: list[int],
) -> list[str]:
    if not text:
        return [""] * len(line_counts)

    text_tokens, spans = tokenize_with_spans(text)
    if not text_tokens:
        return [""] * len(line_counts)

    if sum(part_word_counts) > 0:
        text_token_counts = allocate_units(len(text_tokens), part_word_counts)
    else:
        text_token_counts = allocate_units(len(text_tokens), line_counts)

    boundaries: list[int] = []
    cumulative_tokens = 0
    for count in text_token_counts[:-1]:
        cumulative_tokens += count
        boundaries.append(cumulative_tokens)

    pieces: list[str] = []
    start_token = 0
    for end_token in boundaries + [len(text_tokens)]:
        if start_token >= end_token:
            pieces.append("")
        else:
            start_char = spans[start_token][0]
            end_char = spans[end_token - 1][1]
            pieces.append(text[start_char:end_char].strip())
        start_token = end_token
    return pieces


def build_segment(
    daf: int,
    ref_part: RefPart,
    order: int,
    witness_names: list[str],
    texts: list[str],
    word_rows: list[list[str]],
    vilna_index: int | None,
) -> Segment:
    return Segment(
        daf=daf,
        amud=ref_part.amud,
        ref_label=ref_part.label,
        start_line=ref_part.line_start,
        order=order,
        witness_names=witness_names,
        texts=texts,
        word_rows=word_rows,
        vilna_index=vilna_index,
    )


def split_segment(
    current_daf: int,
    ref_parts: list[RefPart],
    order: int,
    witness_names: list[str],
    texts: list[str],
    word_rows: list[list[str]],
    vilna_index: int | None,
) -> list[Segment]:
    line_counts = [part.line_count for part in ref_parts]
    boundaries = choose_split_boundaries(word_rows, vilna_index, line_counts)
    word_row_parts = split_rows_by_boundaries(word_rows, boundaries)

    text_parts_by_witness = [
        split_text_by_token_counts(
            text=texts[witness_index],
            part_word_counts=[count_nonempty_tokens(rows, witness_index) for rows in word_row_parts],
            line_counts=line_counts,
        )
        for witness_index in range(len(witness_names))
    ]

    split_texts: list[list[str]] = [[] for _ in ref_parts]
    for witness_part_texts in text_parts_by_witness:
        for segment_index, part_text in enumerate(witness_part_texts):
            split_texts[segment_index].append(part_text)

    segments: list[Segment] = []
    daf = current_daf
    previous_part: RefPart | None = None
    for index, ref_part in enumerate(ref_parts):
        if previous_part is not None and previous_part.amud == 'ע"ב' and ref_part.amud == 'ע"א':
            daf += 1
        segments.append(
            build_segment(
                daf=daf,
                ref_part=ref_part,
                order=order + index,
                witness_names=witness_names,
                texts=split_texts[index],
                word_rows=word_row_parts[index],
                vilna_index=vilna_index,
            )
        )
        previous_part = ref_part
    return segments


def build_segments_for_row(
    current_daf: int,
    ref: str,
    order: int,
    witness_names: list[str],
    texts: list[str],
    word_rows: list[list[str]],
    vilna_index: int | None,
) -> list[Segment]:
    ref_strings = [normalize_spaces(part) for part in ref.split(";") if normalize_spaces(part)]
    ref_parts = [parse_ref_part(part) for part in ref_strings]
    if len(ref_parts) >= 2 and all(ref_parts):
        return split_segment(
            current_daf=current_daf,
            ref_parts=[part for part in ref_parts if part is not None],
            order=order,
            witness_names=witness_names,
            texts=texts,
            word_rows=word_rows,
            vilna_index=vilna_index,
        )

    parsed_ref = parse_ref_part(ref)
    if parsed_ref is None:
        warn(f"Could not parse reference {ref!r}; keeping it under ע\"א of daf {current_daf}.")
        parsed_ref = RefPart(amud='ע"א', line_start=0, line_end=0, label=ref)

    return [
        build_segment(
            daf=current_daf,
            ref_part=parsed_ref,
            order=order,
            witness_names=witness_names,
            texts=texts,
            word_rows=word_rows,
            vilna_index=vilna_index,
        )
    ]


def highlight_segment_text(segment: Segment, witness_index: int) -> str:
    text = segment.texts[witness_index]
    if not text:
        return ""
    if segment.vilna_index is None or segment.vilna_index >= len(segment.texts):
        return text
    if witness_index == segment.vilna_index:
        return text

    vilna_text = segment.texts[segment.vilna_index]
    if not vilna_text:
        return text

    word_tokens: list[str] = []
    changed_positions: list[int] = []
    for row in segment.word_rows:
        witness_token = row[witness_index] if witness_index < len(row) else ""
        vilna_token = row[segment.vilna_index] if segment.vilna_index < len(row) else ""
        if not witness_token:
            continue
        token_position = len(word_tokens)
        word_tokens.append(witness_token)
        if normalize_compare_token(witness_token) != normalize_compare_token(vilna_token):
            changed_positions.append(token_position)

    if word_tokens:
        highlighted = apply_word_sheet_highlights(text, word_tokens, changed_positions)
        if highlighted is not None:
            return highlighted

    return apply_diff_highlights(text, vilna_text)


def render_tractate_output(tractate_name: str, segments: list[Segment]) -> str:
    by_daf: dict[int, dict[str, list[Segment]]] = {}
    for segment in segments:
        by_daf.setdefault(segment.daf, {'ע"א': [], 'ע"ב': []})
        by_daf[segment.daf].setdefault(segment.amud, []).append(segment)

    for amud_map in by_daf.values():
        for segment_list in amud_map.values():
            segment_list.sort(key=lambda item: (item.start_line, item.order))

    lines: list[str] = [f"<h1>הכי גרסינן {tractate_name}</h1>", ""]
    for daf in sorted(by_daf):
        for amud, suffix in (('ע"א', "."), ('ע"ב', ":")):
            lines.append(f"<h2>דף {int_to_hebrew(daf)}{suffix}</h2>")
            lines.append("")
            for segment in by_daf[daf].get(amud, []):
                for witness_index, witness_name in enumerate(segment.witness_names):
                    rendered_text = highlight_segment_text(segment, witness_index)
                    if not rendered_text:
                        continue
                    lines.append(f"<b>{witness_name}</b>: {rendered_text}")
                lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def load_segments_from_workbook(path: Path, order_seed: int) -> tuple[list[Segment], int]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    if len(workbook.sheetnames) < 2:
        raise ValueError(f"Workbook has fewer than two sheets: {path}")

    words_sheet_name = choose_sheet_name(workbook.sheetnames, "מילים", 0)
    lines_sheet_name = choose_sheet_name(workbook.sheetnames, "שורות", 1)
    words_sheet = workbook[words_sheet_name]
    lines_sheet = workbook[lines_sheet_name]

    witness_columns = get_witness_columns(lines_sheet)
    witness_names = [name for _column, name in witness_columns]
    try:
        vilna_index = witness_names.index(WILNA_NAME)
    except ValueError:
        vilna_index = None
        warn(f"Workbook {path.name} has no {WILNA_NAME} column.")

    word_rows_by_ref = read_word_rows(words_sheet, witness_columns)
    current_daf = extract_daf_number_from_filename(path)

    segments: list[Segment] = []
    current_order = order_seed
    for row in lines_sheet.iter_rows(min_row=4, values_only=True):
        ref = clean_ref_text(row[0] if row else None)
        if not ref:
            continue

        texts: list[str] = []
        for column_index, _name in witness_columns:
            value = row[column_index - 1] if column_index - 1 < len(row) else None
            texts.append(clean_cell_text(value))

        word_rows = word_rows_by_ref.get(ref, [])
        row_segments = build_segments_for_row(
            current_daf=current_daf,
            ref=ref,
            order=current_order,
            witness_names=witness_names,
            texts=texts,
            word_rows=word_rows,
            vilna_index=vilna_index,
        )
        segments.extend(row_segments)
        current_order += len(row_segments) + 1

    return segments, current_order


def iter_tractate_dirs(source_dir: Path, tractate_filter: str | None) -> list[tuple[str, Path]]:
    candidates: list[tuple[int, str, Path]] = []
    filter_text = normalize_spaces(tractate_filter) if tractate_filter else None

    for child in source_dir.iterdir():
        if not child.is_dir() or child.name == INTRO_DIR_NAME:
            continue
        parsed = parse_tractate_dir(child)
        if parsed is None:
            continue
        order, raw_name = parsed
        tractate_name = normalize_tractate_name(raw_name)
        if filter_text and filter_text not in tractate_name:
            continue
        candidates.append((order, tractate_name, child))

    return [(name, path) for _order, name, path in sorted(candidates)]


def iter_workbooks(tractate_dir: Path) -> list[Path]:
    workbooks: list[tuple[int, Path]] = []
    for path in tractate_dir.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if SPECIAL_FILE_MARKER in path.name:
            continue
        try:
            daf = extract_daf_number_from_filename(path)
        except ValueError as exc:
            warn(str(exc))
            continue
        workbooks.append((daf, path))
    workbooks.sort(key=lambda item: (item[0], item[1].name))
    return [path for _daf, path in workbooks]


def process_tractate(tractate_name: str, tractate_dir: Path, output_dir: Path) -> Path:
    all_segments: list[Segment] = []
    order_seed = 0
    workbooks = iter_workbooks(tractate_dir)
    if not workbooks:
        warn(f"No workbook files found for tractate {tractate_name}.")

    for workbook_path in workbooks:
        try:
            segments, order_seed = load_segments_from_workbook(workbook_path, order_seed)
        except Exception as exc:  # noqa: BLE001
            warn(f"Skipping workbook {workbook_path}: {exc}")
            continue
        all_segments.extend(segments)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"הכי גרסינן {tractate_name}.txt"
    output_text = render_tractate_output(tractate_name, all_segments)
    output_path.write_text(output_text, encoding="utf-8")
    return output_path


def main() -> int:
    args = parse_args()
    source_dir = args.source.resolve()
    output_dir = args.output.resolve()

    if not source_dir.exists():
        warn(f"Source directory does not exist: {source_dir}")
        return 1

    tractates = iter_tractate_dirs(source_dir, args.tractate)
    if not tractates:
        warn("No matching tractate directories were found.")
        return 1

    for tractate_name, tractate_dir in tractates:
        print(f"Processing {tractate_name}...", file=sys.stderr)
        output_path = process_tractate(tractate_name, tractate_dir, output_dir)
        print(f"Wrote {output_path}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
